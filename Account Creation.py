from appium import webdriver
from appium.webdriver.common.mobileby import MobileBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from appium.webdriver.common.touch_action import TouchAction
from selenium.webdriver.common.touch_actions import TouchActions
from selenium.webdriver.common.action_chains import ActionChains
desired_cap = {
    # Set your access credentials
    "browserstack.user": "jayanth_AZJzjH",
    "browserstack.key": "CbunYNkQWfpY6LCMqSTU",

    # Set URL of the application under test
    "app": "bs://bb84e0811789aece00d7a3655602fdd5a01df662",

    # Specify device and os_version for testing
    "device": "Google Pixel 3",
    "os_version": "9.0",

    # Set other BrowserStack capabilities
    "project": "First Python project",
    "build": "Python Android",
    "name": "first_test"
}
userName = 'jayanth_AZJzjH'
accessKey = 'CbunYNkQWfpY6LCMqSTU'
# Initialize the remote Webdriver using BrowserStack remote URL
# and desired capabilities defined above
driver = webdriver.Remote("https://" + userName + ":" + accessKey + "@hub-cloud.browserstack.com/wd/hub", desired_cap)
usernames = ['darklord3244,adjisd23']
wb = openpyxl.load_workbook('E:\\Python\\Pycharm\\Projects\\Final.xlsx')
sheet = wb.active
final_usr = 0
final_mail = 0
final_mail1 = 0
final_mail2 = 0
final_mail3 = 0
final_mail4 = 0
u =0
for q in range(5,72):
    u = q
    cat = str(u)
    final_mail1 = sheet.cell(row=u,column=1).value
    final_mail2 = sheet.cell(row=u, column=2).value
    final_mail3 = sheet.cell(row=u, column=3).value
    final_mail4 = int(final_mail3)
    final_usr = str(final_mail4)
    final_mail = final_mail1+final_mail2+final_usr
    emails =final_mail+'@mail.com'
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/primary_button').click()
    time.sleep(2)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/btn_next').click()
    time.sleep(2)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/btn_next').click()
    time.sleep(2)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/btn_get_started').click()
    time.sleep(2)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/btn_signup').click()
    time.sleep(2)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/edit_text_email').click()
    email = driver.find_element_by_id('sisinc.com.sis:id/edit_text_email')
    email.send_keys(emails)
    time.sleep(1)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/edit_text_username').click()
    usrname = driver.find_element_by_id('sisinc.com.sis:id/edit_text_username')
    usrname.send_keys(usernames[q])
    time.sleep(1)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/edit_text_password').click()
    password = driver.find_element_by_id('sisinc.com.sis:id/edit_text_password')
    password.send_keys("Am1zomem_123",)
    time.sleep(1)
    driver.back()
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/checkbox_terms_conditions').click()
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/btn_register').click()
    time.sleep(5)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/ld_btn_yes').click()
    time.sleep(10)
    imga = '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout/android.widget.RelativeLayout/androidx.recyclerview.widget.RecyclerView/android.widget.RelativeLayout[1]/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.RelativeLayout/android.widget.ImageView'
    driver.implicitly_wait(30)
    driver.find_element_by_xpath(imga).click()
    imga1 = '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout/android.widget.RelativeLayout/androidx.recyclerview.widget.RecyclerView/android.widget.RelativeLayout[2]/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.RelativeLayout/android.widget.ImageView'
    driver.implicitly_wait(30)
    driver.find_element_by_xpath(imga1).click()
    imga2  = '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout/android.widget.RelativeLayout/androidx.recyclerview.widget.RecyclerView/android.widget.RelativeLayout[3]/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.RelativeLayout/android.widget.ImageView'
    driver.implicitly_wait(30)
    driver.find_element_by_xpath(imga2).click()
    imga3  = '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout/android.widget.RelativeLayout/androidx.recyclerview.widget.RecyclerView/android.widget.RelativeLayout[4]/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.RelativeLayout/android.widget.ImageView'
    driver.implicitly_wait(30)
    driver.find_element_by_xpath(imga3).click()
    imga4  = '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout/android.widget.RelativeLayout/androidx.recyclerview.widget.RecyclerView/android.widget.RelativeLayout[5]/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.RelativeLayout/android.widget.ImageView'
    driver.implicitly_wait(30)
    driver.find_element_by_xpath(imga4).click()
    imga5  = '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout/android.widget.RelativeLayout/androidx.recyclerview.widget.RecyclerView/android.widget.RelativeLayout[6]/android.widget.FrameLayout/android.widget.RelativeLayout/android.widget.RelativeLayout/android.widget.ImageView'
    driver.implicitly_wait(30)
    driver.find_element_by_xpath(imga5).click()
    save = 'sisinc.com.sis:id/save_cats'
    driver.implicitly_wait(30)
    driver.find_element_by_id(save).click()

    time.sleep(3)
    search_img = 'sisinc.com.sis:id/img_search'
    driver.implicitly_wait(30)
    driver.find_element_by_id(search_img).click()
    search_img1 = 'sisinc.com.sis:id/editText3'
    driver.implicitly_wait(30)
    search_img2 = driver.find_element_by_id(search_img1)
    search_img2.send_keys('NikhillxD')
    driver.back()
    search_im3 = '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.RelativeLayout/androidx.viewpager.widget.ViewPager/android.widget.RelativeLayout/android.widget.ListView/android.widget.RelativeLayout[1]/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.ImageView'
    driver.implicitly_wait(30)
    driver.find_element_by_xpath(search_im3).click()
    time.sleep(5)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/follow').click()
    time.sleep(1)
    TouchAction(driver).press(x=1028, y=1493).wait(1000).move_to(x=1033, y=470).release().perform()
    a = 0
    for y in range(1,9):
        a = y
        b = str(a)
        search_im4 = '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.ScrollView/androidx.viewpager.widget.ViewPager/android.widget.RelativeLayout/android.view.ViewGroup/androidx.recyclerview.widget.RecyclerView/android.view.ViewGroup['+b+']/android.widget.ImageView'
        driver.implicitly_wait(30)
        driver.find_element_by_xpath(search_im4).click()
        time.sleep(5)
        driver.find_element_by_id('sisinc.com.sis:id/button1').click()
        driver.back()

    time.sleep(5)
    driver.back()
    time.sleep(1)
    driver.back()
    time.sleep(1)
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/action_item5').click()
    driver.implicitly_wait(30)
    driver.find_element_by_id('sisinc.com.sis:id/appsetting').click()
    driver.implicitly_wait(30)
    driver.find_element_by_xpath('/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.LinearLayout[11]/android.widget.ImageView').click()
    time.sleep(10)
for x in range(1,10):
    value = driver.orientation
    time.sleep(50)